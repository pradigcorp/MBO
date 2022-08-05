from django.shortcuts import render
from django.contrib.auth import authenticate,logout#,login
from django.contrib.auth import login as auth_login
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from .models import GOAL22,CPA22
from .forms import GOAL22Q1Form,CPA22CForm,CPA22AForm
import os
#import xlwt
#import xlrd
#from xlutils.copy import copy
import XlsxWriter
import io
import openpyxl as op

# Create your views here.

User = get_user_model()

BB1='①役割認識・責任行動'
BB2='組織内で果たすべき職務責任を理解し意思決定をタイムリーに行っている'
BB3='問題が発生した時は他に責任転嫁せず自らが責任をとるという態度で行動している'
BB4='結果や自己の言動に対して責任回避・転嫁をしない'
BB5='②対人折衝力'
BB6='利用者やスタッフに対して感情的にならず誠実に対応し説得力に富んだ話し方で接している'
BB7='相手の主張に耳を傾けるとともに自己の考えを相手が納得するまで粘り強く説明している'
BB8='セクショナリズムを避けながら内外との調整を行い仕事の計画・基盤づくり等の前準備をしている'
BB9='③戦略的思考能力'
BB10='常に論理的な思考を心がけ問題が生じた時には表面化した問題ではなくその根本原因を把握しようと努めている'
BB11='目の前の業務だけではなく常に将来のありたい姿を見据え現状とのギャップを埋めるための業務をしている'
BB12='不透明性の高い事象に対しても仮説を立てながら取り組み前進することができる'
BB13='④知識の習得'
BB14='職務に関する新しい知識や情報に強い関心を持ち日頃から広く情報収集している'
BB15='知識・情報を業務の改善・改革や日常の職務遂行に活用している'
BB16='業界動向、国や自治体の政策、担当業務の最新情報を常に把握している'
BB17='⑤リーダーシップ'
BB18='会社の理念・方針について理解し所属する組織メンバーと共有しその体現に努めている'
BB19='部下・後輩の能力開発の重要性を理解し積極的・具体的にそのサポートをしている'
BB20='所属する組織が適切かつ迅速に業務を遂行できるよう権限の移譲がなされている'

def home(request):
    form = 'hello'
    return render(request, 'MBO/home.html', {'form': form})

def login(request):
    if request.method == 'POST':
        ID = request.POST.get('userid')
        Pass = request.POST.get('password')
        user = authenticate(username=ID, password=Pass)
        if user:
            if user.is_active:
                auth_login(request,user)
                task = get_object_or_404(GOAL22, user=request.user)
                task.GOAL22A1='課の売り上げ目標10億円の達成'
                task.GOAL22B1='課の新規顧客50社を開拓'
                task.GOAL22C1='チーム員の能力向上の為にコンピテンシー評価を実施し必要な研修・アドバイス・OJTの機会を用意する'
                task.GOAL22D1='営業レポートのクラウドウェブ化によるDX'
                task.GOAL22E1='課の風通しの良い組織風土の醸成'
                task.GOAL22AP=30
                task.GOAL22BP=30
                task.GOAL22CP=20
                task.GOAL22DP=10
                task.GOAL22EP=10
                task.GOAL22Q1=0
                task.GOAL22Q2=0
                task.GOAL22Q3=0
                task.save()
                task2 = get_object_or_404(CPA22, user=request.user)
                task2.CPA22A1C=1
                task2.CPA22A1A=1
                task2.CPA22A1K='権限移譲の不足により組織としての意思決定が遅れることがある'
                task2.CPA22A1P='アサーティブネスにつき一度受講した上で権限を一度見直し部下に一部移譲することを検討する'
                task2.CPA22A2C=3
                task2.CPA22A2A=3
                task2.CPA22A2K='常に自責を意識し無意識に他責思考になっていないか意識的に気を付けている'
                task2.CPA22A2P='引き続き自責思考を継続する'
                task2.CPA22A3C=2
                task2.CPA22A3A=1
                task2.CPA22A3K='結果が達成できなかった時に責任回避する時がある'
                task2.CPA22A3P='結果が達成できなかった時には言い訳ではなく説明責任を果たすことに集中する'
                task2.CPA22B1C=2
                task2.CPA22B1A=1
                task2.CPA22B1K='時に印象論に偏り論理的でないことがある'
                task2.CPA22B1P='ロジカルシンキングを受講した上で論理的思考を会得する'
                task2.CPA22B2C=3
                task2.CPA22B2A=3
                task2.CPA22B2K='強みのひとつで聞くことを知っており粘り強さもある'
                task2.CPA22B2P='強みとして自身で認識するとともに業務への落とし込みをより意識し成果へと繋げる'
                task2.CPA22B3C=3
                task2.CPA22B3A=2
                task2.CPA22B3K='強みのひとつで特に社内根回しは秀逸'
                task2.CPA22B3P='一方で社外との折衝における準備は改善余地あり、相手の状況をよく検討した上で提案を準備する'
                task2.CPA22C1C=1
                task2.CPA22C1A=1
                task2.CPA22C1K='時に表層の問題に囚われ根本原因を見落とすことがある'
                task2.CPA22C1P='5　WHYを常にこころがけ根本原因を見極める力・習性を会得する'
                task2.CPA22C2C=3
                task2.CPA22C2A=2
                task2.CPA22C2K='よく組織の戦略を理解しておりその戦術の遂行をしっかり行っている'
                task2.CPA22C2P='現状は良きフォロワーであるが自らも意識的に組織のありたい姿を描き戦略に落とし込む訓練をする'
                task2.CPA22C3C=1
                task2.CPA22C3A=1
                task2.CPA22C3K='過去例の無い新規事業や新規問題に対峙した時に対応・思考が止まってしまうことがある'
                task2.CPA22C3P='仮説検証による事業推進につき一度学習しＯＪＴを通じた習得を目指す'
                task2.CPA22D1C=3
                task2.CPA22D1A=3
                task2.CPA22D1K='常に学習意欲を高く持ち様々なことに興味を有していると思う'
                task2.CPA22D1P='強みのひとつで継続していく'
                task2.CPA22D2C=2
                task2.CPA22D2A=2
                task2.CPA22D2K='知識量は強みのひとつである一方業務への落とし込みが不足している'
                task2.CPA22D2P='知識量が自身の強みであることを認知し常にそれを業務・成果に落とし込むことを意識する'
                task2.CPA22D3C=1
                task2.CPA22D3A=1
                task2.CPA22D3K='業界におけるコネクションが不足していることから時に情報アップデートが遅れている'
                task2.CPA22D3P='意識的に業界紙や集会に触れる機会を持つ'
                task2.CPA22E1C=2
                task2.CPA22E1A=2
                task2.CPA22E1K='よく会社の理念・方針を理解している'
                task2.CPA22E1P='課長としてチームメンバーの理解深化にさらに努めていく'
                task2.CPA22E2C=3
                task2.CPA22E2A=2
                task2.CPA22E2K='チームとして戦う力の重要性をよく理解しており常にサポートを意識している'
                task2.CPA22E2P='より高度なチーム員の能力開発のために他部署の参考例を学ぶ'
                task2.CPA22E3C=1
                task2.CPA22E3A=1
                task2.CPA22E3K='権限移譲の不足により組織としての意思決定が遅れることがある'
                task2.CPA22E3P='アサーティブネスにつき一度受講した上で権限を一度見直し部下に一部移譲することを検討する'
                task2.CPA22C=1
                task2.save()
                data='nothing'
                THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
                my_file = os.path.join(THIS_FOLDER,'goals.txt')
                f = open(my_file,'r', encoding='UTF-8')
                data = f.read()
                data = data.replace('\n','^')
                data = data.split('^')
                for i in range(4):
                    j=i+3
                    try:
                        name=User.objects.get(id=j)
                        add_user = User.objects.filter(username=name.username).first()
                        DD = GOAL22()
                        DD.user = add_user
                        DD.save()
                        EE = CPA22()
                        EE.user = add_user
                        EE.save()
                    except:
                        yuto=1
                    task = get_object_or_404(GOAL22, user=j)
                    task.GOAL22A1=data[i*42]
                    task.GOAL22B1=data[i*42+1]
                    task.GOAL22C1=data[i*42+2]
                    task.GOAL22D1=data[i*42+3]
                    task.GOAL22E1=data[i*42+4]
                    task.GOAL22AP=data[i*42+7]
                    task.GOAL22BP=data[i*42+8]
                    task.GOAL22CP=data[i*42+9]
                    task.GOAL22DP=data[i*42+10]
                    task.GOAL22EP=data[i*42+11]
                    task.GOAL22FP=data[i*42+12]
                    task.GOAL22GP=data[i*42+13]
                    task.GOAL22A2=data[i*42+14]
                    task.GOAL22B2=data[i*42+15]
                    task.GOAL22C2=data[i*42+16]
                    task.GOAL22D2=data[i*42+17]
                    task.GOAL22E2=data[i*42+18]
                    task.GOAL22A3=data[i*42+21]
                    task.GOAL22B3=data[i*42+22]
                    task.GOAL22C3=data[i*42+23]
                    task.GOAL22D3=data[i*42+24]
                    task.GOAL22E3=data[i*42+25]
                    task.GOAL22AR=data[i*42+28]
                    task.GOAL22BR=data[i*42+29]
                    task.GOAL22CR=data[i*42+30]
                    task.GOAL22DR=data[i*42+31]
                    task.GOAL22ER=data[i*42+32]
                    task.GOAL22FR=data[i*42+33]
                    task.GOAL22GR=data[i*42+34]
                    task.GOAL22AJ=data[i*42+35]
                    task.GOAL22BJ=data[i*42+36]
                    task.GOAL22CJ=data[i*42+37]
                    task.GOAL22DJ=data[i*42+38]
                    task.GOAL22EJ=data[i*42+39]
                    task.GOAL22Q1=1
                    task.GOAL22Q1A=0
                    if data[i*42+5] != 'A':
                        task.GOAL22F1=data[i*42+5]
                        task.GOAL22G1=data[i*42+6]
                        task.GOAL22FJ=data[i*42+40]
                        task.GOAL22GJ=data[i*42+41]
                        task.GOAL22F2=data[i*42+19]
                        task.GOAL22G2=data[i*42+20]
                        task.GOAL22F3=data[i*42+26]
                        task.GOAL22G3=data[i*42+27]
                        task.GOAL22F1=data[i*42+5]
                        task.GOAL22G1=data[i*42+6]
                        task.GOAL22FJ=data[i*42+40]
                        task.GOAL22GJ=data[i*42+41]
                        task.GOAL22F2=data[i*42+19]
                        task.GOAL22G2=data[i*42+20]
                        task.GOAL22F3=data[i*42+26]
                        task.GOAL22G3=data[i*42+27]
                    else:
                        task.GOAL22F1=''
                        task.GOAL22G1=''
                        task.GOAL22FJ=''
                        task.GOAL22GJ=''
                        task.GOAL22F2=''
                        task.GOAL22G2=''
                        task.GOAL22F3=''
                        task.GOAL22G3=''
                        task.GOAL22F1=''
                        task.GOAL22G1=''
                        task.GOAL22FJ=''
                        task.GOAL22GJ=''
                        task.GOAL22F2=''
                        task.GOAL22G2=''
                        task.GOAL22F3=''
                        task.GOAL22G3=''
                    task.save()

                my_file2 = os.path.join(THIS_FOLDER,'cpas.txt')
                f2 = open(my_file2,'r', encoding='UTF-8')
                data2 = f2.read()
                data2 = data2.replace('\n','^')
                data2 = data2.split('^')
                for i in range(4):
                    j=i+3
                    task3 = get_object_or_404(CPA22, user=j)
                    task3.CPA22A1C=data2[i*60]
                    task3.CPA22A1A=data2[i*60+1]
                    task3.CPA22A1K=data2[i*60+2]
                    task3.CPA22A1P=data2[i*60+3]
                    task3.CPA22A2C=data2[i*60+4]
                    task3.CPA22A2A=data2[i*60+5]
                    task3.CPA22A2K=data2[i*60+6]
                    task3.CPA22A2P=data2[i*60+7]
                    task3.CPA22A3C=data2[i*60+8]
                    task3.CPA22A3A=data2[i*60+9]
                    task3.CPA22A3K=data2[i*60+10]
                    task3.CPA22A3P=data2[i*60+11]
                    task3.CPA22B1C=data2[i*60+12]
                    task3.CPA22B1A=data2[i*60+13]
                    task3.CPA22B1K=data2[i*60+14]
                    task3.CPA22B1P=data2[i*60+15]
                    task3.CPA22B2C=data2[i*60+16]
                    task3.CPA22B2A=data2[i*60+17]
                    task3.CPA22B2K=data2[i*60+18]
                    task3.CPA22B2P=data2[i*60+19]
                    task3.CPA22B3C=data2[i*60+20]
                    task3.CPA22B3A=data2[i*60+21]
                    task3.CPA22B3K=data2[i*60+22]
                    task3.CPA22B3P=data2[i*60+23]
                    task3.CPA22C1C=data2[i*60+24]
                    task3.CPA22C1A=data2[i*60+25]
                    task3.CPA22C1K=data2[i*60+26]
                    task3.CPA22C1P=data2[i*60+27]
                    task3.CPA22C2C=data2[i*60+28]
                    task3.CPA22C2A=data2[i*60+29]
                    task3.CPA22C2K=data2[i*60+30]
                    task3.CPA22C2P=data2[i*60+31]
                    task3.CPA22C3C=data2[i*60+32]
                    task3.CPA22C3A=data2[i*60+33]
                    task3.CPA22C3K=data2[i*60+34]
                    task3.CPA22C3P=data2[i*60+35]
                    task3.CPA22D1C=data2[i*60+36]
                    task3.CPA22D1A=data2[i*60+37]
                    task3.CPA22D1K=data2[i*60+38]
                    task3.CPA22D1P=data2[i*60+39]
                    task3.CPA22D2C=data2[i*60+40]
                    task3.CPA22D2A=data2[i*60+41]
                    task3.CPA22D2K=data2[i*60+42]
                    task3.CPA22D2P=data2[i*60+43]
                    task3.CPA22D3C=data2[i*60+44]
                    task3.CPA22D3A=data2[i*60+45]
                    task3.CPA22D3K=data2[i*60+46]
                    task3.CPA22D3P=data2[i*60+47]
                    task3.CPA22E1C=data2[i*60+48]
                    task3.CPA22E1A=data2[i*60+49]
                    task3.CPA22E1K=data2[i*60+50]
                    task3.CPA22E1P=data2[i*60+51]
                    task3.CPA22E2C=data2[i*60+52]
                    task3.CPA22E2A=data2[i*60+53]
                    task3.CPA22E2K=data2[i*60+54]
                    task3.CPA22E2P=data2[i*60+55]
                    task3.CPA22E3C=data2[i*60+56]
                    task3.CPA22E3A=data2[i*60+57]
                    task3.CPA22E3K=data2[i*60+58]
                    task3.CPA22E3P=data2[i*60+59]
                    task3.CPA22C=1
                    task3.CPA22A=0
                    task3.save()
                return HttpResponseRedirect(reverse('sample'))
            else:
                return HttpResponse("アカウントが有効ではありません")
        else:
            return HttpResponse("ログインIDかパスワードに誤りがあります")
    else:
        return render(request, 'MBO/login.html')

def sample(request):
    params = {
        "FN" : request.user.first_name,
        "LN" : request.user.last_name,
    }
    return render(request, 'MBO/sample.html', params)

def GOAL(request):

    AA=GOAL22.objects.values_list('GOAL22Q1').get(user=request.user)

    if int(AA[0]) == 0:
        sbmt=0
    else:
        sbmt=1

    task = get_object_or_404(GOAL22, user=request.user)
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = XlsxWriter.Workbook(output)
            ws = book.add_worksheet('goalsetting')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'期初目標設定', format)
            ws.write(1,2,'ウェイト%', format)
            ws.write(1,3,'期中レビュー', format)
            ws.write(1,4,'期末レビュー', format)
            ws.write(1,5,'評価点数', format)
            ws.write(1,6,'評定者コメント', format)
            ws.write(2,1,task.GOAL22A1, format)
            ws.write(2,2,task.GOAL22AP, format)
            ws.write(2,3,task.GOAL22A2, format)
            ws.write(2,4,task.GOAL22A3, format)
            ws.write(2,5,task.GOAL22AR, format)
            ws.write(2,6,task.GOAL22AJ, format)
            ws.write(3,1,task.GOAL22B1, format)
            ws.write(3,2,task.GOAL22BP, format)
            ws.write(3,3,task.GOAL22B2, format)
            ws.write(3,4,task.GOAL22B3, format)
            ws.write(3,5,task.GOAL22BR, format)
            ws.write(4,1,task.GOAL22C1, format)
            ws.write(4,2,task.GOAL22CP, format)
            ws.write(4,3,task.GOAL22C2, format)
            ws.write(4,4,task.GOAL22C3, format)
            ws.write(4,5,task.GOAL22CR, format)
            ws.write(5,1,task.GOAL22D1, format)
            ws.write(5,2,task.GOAL22DP, format)
            ws.write(5,3,task.GOAL22D2, format)
            ws.write(5,4,task.GOAL22D3, format)
            ws.write(5,5,task.GOAL22DR, format)
            ws.write(6,1,task.GOAL22E1, format)
            ws.write(6,2,task.GOAL22EP, format)
            ws.write(6,3,task.GOAL22E2, format)
            ws.write(6,4,task.GOAL22E3, format)
            ws.write(6,5,task.GOAL22ER, format)
            ws.write(7,1,task.GOAL22F1, format)
            ws.write(7,2,task.GOAL22FP, format)
            ws.write(7,3,task.GOAL22F2, format)
            ws.write(7,4,task.GOAL22F3, format)
            ws.write(7,5,task.GOAL22FR, format)
            ws.write(8,1,task.GOAL22G1, format)
            ws.write(8,2,task.GOAL22GP, format)
            ws.write(8,3,task.GOAL22G2, format)
            ws.write(8,4,task.GOAL22G3, format)
            ws.write(8,5,task.GOAL22GR, format)
            ws.write(3,6,task.GOAL22BJ, format)
            ws.write(4,6,task.GOAL22CJ, format)
            ws.write(5,6,task.GOAL22DJ, format)
            ws.write(6,6,task.GOAL22EJ, format)
            ws.write(7,6,task.GOAL22FJ, format)
            ws.write(8,6,task.GOAL22GJ, format)
            ws.set_column('B:B',33)
            ws.set_column('C:C',11)
            ws.set_column('D:E',33)
            ws.set_column('F:F',11)
            ws.set_column('G:G',33)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(5,50)
            ws.set_row(6,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            book.close()
            output.seek(0)
            filename = 'goalsetting.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response
        elif "send" in request.POST:        
            sbmt = 1
            task.GOAL22Q1 = sbmt
            task.save()

    params = {
        "UserID":request.user,
        "data1":GOAL22.objects.values_list('GOAL22A1','GOAL22B1','GOAL22C1','GOAL22D1','GOAL22E1','GOAL22F1','GOAL22G1').get(user=request.user),
        "data2":GOAL22.objects.values_list('GOAL22AP','GOAL22BP','GOAL22CP','GOAL22DP','GOAL22EP','GOAL22FP','GOAL22GP').get(user=request.user),
        "data3":GOAL22.objects.values_list('GOAL22A2','GOAL22B2','GOAL22C2','GOAL22D2','GOAL22E2','GOAL22F2','GOAL22G2').get(user=request.user),
        "data4":GOAL22.objects.values_list('GOAL22A3','GOAL22B3','GOAL22C3','GOAL22D3','GOAL22E3','GOAL22F3','GOAL22G3').get(user=request.user),
        "data5":GOAL22.objects.values_list('GOAL22AR','GOAL22BR','GOAL22CR','GOAL22DR','GOAL22ER','GOAL22FR','GOAL22GR').get(user=request.user),
        "data6":GOAL22.objects.values_list('GOAL22AJ','GOAL22BJ','GOAL22CJ','GOAL22DJ','GOAL22EJ','GOAL22FJ','GOAL22GJ').get(user=request.user),
        "sbmt":sbmt,
        "FN":request.user.first_name,
        "LN":request.user.last_name,
        }
    return render(request, 'MBO/GOAL.html', params)


def GOAL2(request, name):

    object_list = User.objects.all()
    for i in object_list:
        if i.username == name:
            ID = i.id
            FN = i.first_name
            LN = i.last_name
    
    AA=GOAL22.objects.values_list('GOAL22Q1').get(user=ID)
    CC=GOAL22.objects.values_list('GOAL22Q1A').get(user=ID)
    if int(AA[0]) == 0:
        sbmt=0
    else:
        if int(CC[0]) == 0:
            sbmt = 1
        else:
            sbmt = 2

    task = get_object_or_404(GOAL22, user=ID)
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = XlsxWriter.Workbook(output)
            ws = book.add_worksheet('goalsetting')
            format = book.add_format({'border':1})          
            ws.write(1,1,'期初目標設定', format)
            ws.write(1,2,'ウェイト%', format)
            ws.write(1,3,'期中レビュー', format)
            ws.write(1,4,'期末レビュー', format)
            ws.write(1,5,'評価点数', format)
            ws.write(2,1,task.GOAL22A1, format)
            ws.write(2,2,task.GOAL22AP, format)
            ws.write(2,3,task.GOAL22A2, format)
            ws.write(2,4,task.GOAL22A3, format)
            ws.write(2,5,task.GOAL22AR, format)
            ws.write(3,1,task.GOAL22B1, format)
            ws.write(3,2,task.GOAL22BP, format)
            ws.write(3,3,task.GOAL22B2, format)
            ws.write(3,4,task.GOAL22B3, format)
            ws.write(3,5,task.GOAL22BR, format)
            ws.write(4,1,task.GOAL22C1, format)
            ws.write(4,2,task.GOAL22CP, format)
            ws.write(4,3,task.GOAL22C2, format)
            ws.write(4,4,task.GOAL22C3, format)
            ws.write(4,5,task.GOAL22CR, format)
            ws.write(5,1,task.GOAL22D1, format)
            ws.write(5,2,task.GOAL22DP, format)
            ws.write(5,3,task.GOAL22D2, format)
            ws.write(5,4,task.GOAL22D3, format)
            ws.write(5,5,task.GOAL22DR, format)
            ws.write(6,1,task.GOAL22E1, format)
            ws.write(6,2,task.GOAL22EP, format)
            ws.write(6,3,task.GOAL22E2, format)
            ws.write(6,4,task.GOAL22E3, format)
            ws.write(6,5,task.GOAL22ER, format)
            ws.write(7,1,task.GOAL22F1, format)
            ws.write(7,2,task.GOAL22FP, format)
            ws.write(7,3,task.GOAL22F2, format)
            ws.write(7,4,task.GOAL22F3, format)
            ws.write(7,5,task.GOAL22FR, format)
            ws.write(8,1,task.GOAL22G1, format)
            ws.write(8,2,task.GOAL22GP, format)
            ws.write(8,3,task.GOAL22G2, format)
            ws.write(8,4,task.GOAL22G3, format)
            ws.write(8,5,task.GOAL22GR, format)
            ws.set_column('B:B',50)
            ws.set_column('C:C',12)
            ws.set_column('D:E',50)
            ws.set_column('F:F',12)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(5,50)
            ws.set_row(6,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            book.close()
            output.seek(0)
            filename = 'goalsetting.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response
        elif "send" in request.POST:        
            sbmt = 2
            task.GOAL22Q1A = 1
            task.save()
        elif "back" in request.POST:        
            sbmt = 0
            task.GOAL22Q1 = sbmt
            task.save()        

    params = {
        "data1":GOAL22.objects.values_list('GOAL22A1','GOAL22B1','GOAL22C1','GOAL22D1','GOAL22E1','GOAL22F1','GOAL22G1').get(user=ID),
        "data2":GOAL22.objects.values_list('GOAL22AP','GOAL22BP','GOAL22CP','GOAL22DP','GOAL22EP','GOAL22FP','GOAL22GP').get(user=ID),
        "data3":GOAL22.objects.values_list('GOAL22A2','GOAL22B2','GOAL22C2','GOAL22D2','GOAL22E2','GOAL22F2','GOAL22G2').get(user=ID),
        "data4":GOAL22.objects.values_list('GOAL22A3','GOAL22B3','GOAL22C3','GOAL22D3','GOAL22E3','GOAL22F3','GOAL22G3').get(user=ID),
        "data5":GOAL22.objects.values_list('GOAL22AR','GOAL22BR','GOAL22CR','GOAL22DR','GOAL22ER','GOAL22FR','GOAL22GR').get(user=ID),
        "data6":GOAL22.objects.values_list('GOAL22AJ','GOAL22BJ','GOAL22CJ','GOAL22DJ','GOAL22EJ','GOAL22FJ','GOAL22GJ').get(user=ID),
        "sbmt":sbmt,
        "EMP" : name,
        "FN" : FN,
        "LN" : LN,
        "FN2" : request.user.first_name,
        "LN2" : request.user.last_name,
        }
    return render(request, 'MBO/GOAL2.html', params)



def GOAL3(request):

    object_list = User.objects.all()

    task = get_object_or_404(GOAL22, user=request.user)
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = XlsxWriter.Workbook(output)
            ws = book.add_worksheet('goalsetting')
            format = book.add_format({'border':1})
            format.set_text_wrap()         
            ws.write(1,1,'期初目標設定', format)
            ws.write(1,2,'ウェイト%', format)
            ws.write(1,3,'期中レビュー', format)
            ws.write(1,4,'期末レビュー', format)
            ws.write(1,5,'評価点数', format)
            ws.write(1,6,'評定者コメント', format)
            ws.write(2,1,task.GOAL22A1, format)
            ws.write(2,2,task.GOAL22AP, format)
            ws.write(2,3,task.GOAL22A2, format)
            ws.write(2,4,task.GOAL22A3, format)
            ws.write(2,5,task.GOAL22AR, format)
            ws.write(2,6,task.GOAL22AJ, format)
            ws.write(3,1,task.GOAL22B1, format)
            ws.write(3,2,task.GOAL22BP, format)
            ws.write(3,3,task.GOAL22B2, format)
            ws.write(3,4,task.GOAL22B3, format)
            ws.write(3,5,task.GOAL22BR, format)
            ws.write(4,1,task.GOAL22C1, format)
            ws.write(4,2,task.GOAL22CP, format)
            ws.write(4,3,task.GOAL22C2, format)
            ws.write(4,4,task.GOAL22C3, format)
            ws.write(4,5,task.GOAL22CR, format)
            ws.write(5,1,task.GOAL22D1, format)
            ws.write(5,2,task.GOAL22DP, format)
            ws.write(5,3,task.GOAL22D2, format)
            ws.write(5,4,task.GOAL22D3, format)
            ws.write(5,5,task.GOAL22DR, format)
            ws.write(6,1,task.GOAL22E1, format)
            ws.write(6,2,task.GOAL22EP, format)
            ws.write(6,3,task.GOAL22E2, format)
            ws.write(6,4,task.GOAL22E3, format)
            ws.write(6,5,task.GOAL22ER, format)
            ws.write(7,1,task.GOAL22F1, format)
            ws.write(7,2,task.GOAL22FP, format)
            ws.write(7,3,task.GOAL22F2, format)
            ws.write(7,4,task.GOAL22F3, format)
            ws.write(7,5,task.GOAL22FR, format)
            ws.write(8,1,task.GOAL22G1, format)
            ws.write(8,2,task.GOAL22GP, format)
            ws.write(8,3,task.GOAL22G2, format)
            ws.write(8,4,task.GOAL22G3, format)
            ws.write(8,5,task.GOAL22GR, format)
            ws.write(3,6,task.GOAL22BJ, format)
            ws.write(4,6,task.GOAL22CJ, format)
            ws.write(5,6,task.GOAL22DJ, format)
            ws.write(6,6,task.GOAL22EJ, format)
            ws.write(7,6,task.GOAL22FJ, format)
            ws.write(8,6,task.GOAL22GJ, format)
            ws.set_column('B:B',33)
            ws.set_column('C:C',11)
            ws.set_column('D:E',33)
            ws.set_column('F:F',11)
            ws.set_column('G:G',33)
            ws.set_row(2,50)
            ws.set_row(3,50)
            ws.set_row(4,50)
            ws.set_row(5,50)
            ws.set_row(6,50)
            ws.set_row(7,50)
            ws.set_row(8,50)
            book.close()
            output.seek(0)
            filename = 'goalsetting.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response
        elif "send" in request.POST:        
            sbmt = 1
            task.GOAL22Q1 = sbmt
            task.save()

    params = {
        "data1":object_list,
        }
    return render(request, 'MBO/GOAL3.html', params)


def CPA(request):

    AA=CPA22.objects.values_list('CPA22C').get(user=request.user)
    if int(AA[0]) == 0:
        sbmt=0
    else:
        sbmt=1
          
    task = get_object_or_404(CPA22, user=request.user)
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = XlsxWriter.Workbook(output)
            ws = book.add_worksheet('competency')
            format = book.add_format({'border':1})
            format.set_text_wrap()          
            ws.merge_range('B3:B5',BB1, format)
            ws.write(2,2,BB2, format)
            ws.write(3,2,BB3, format)
            ws.write(4,2,BB4, format)
            ws.write(2,3,task.CPA22A1C, format)
            ws.write(3,3,task.CPA22A2C, format)
            ws.write(4,3,task.CPA22A3C, format)
            ws.write(2,4,task.CPA22A1A, format)
            ws.write(3,4,task.CPA22A2A, format)
            ws.write(4,4,task.CPA22A3A, format)
            ws.write(2,5,task.CPA22A1K, format)
            ws.write(3,5,task.CPA22A2K, format)
            ws.write(4,5,task.CPA22A3K, format)
            ws.write(2,6,task.CPA22A1P, format)
            ws.write(3,6,task.CPA22A2P, format)
            ws.write(4,6,task.CPA22A3P, format)
            ws.merge_range('B6:B8',BB5, format)
            ws.write(5,2,BB6, format)
            ws.write(6,2,BB7, format)
            ws.write(7,2,BB8, format)
            ws.write(5,3,task.CPA22B1C, format)
            ws.write(6,3,task.CPA22B2C, format)
            ws.write(7,3,task.CPA22B3C, format)
            ws.write(5,4,task.CPA22B1A, format)
            ws.write(6,4,task.CPA22B2A, format)
            ws.write(7,4,task.CPA22B3A, format)
            ws.write(5,5,task.CPA22B1K, format)
            ws.write(6,5,task.CPA22B2K, format)
            ws.write(7,5,task.CPA22B3K, format)
            ws.write(5,6,task.CPA22B1P, format)
            ws.write(6,6,task.CPA22B2P, format)
            ws.write(7,6,task.CPA22B3P, format)
            ws.merge_range('B9:B11',BB9, format)
            ws.write(8,2,BB10, format)
            ws.write(9,2,BB11, format)
            ws.write(10,2,BB12, format)
            ws.write(8,3,task.CPA22C1C, format)
            ws.write(9,3,task.CPA22C2C, format)
            ws.write(10,3,task.CPA22C3C, format)
            ws.write(8,4,task.CPA22C1A, format)
            ws.write(9,4,task.CPA22C2A, format)
            ws.write(10,4,task.CPA22C3A, format)
            ws.write(8,5,task.CPA22C1K, format)
            ws.write(9,5,task.CPA22C2K, format)
            ws.write(10,5,task.CPA22C3K, format)
            ws.write(8,6,task.CPA22C1P, format)
            ws.write(9,6,task.CPA22C2P, format)
            ws.write(10,6,task.CPA22C3P, format)
            ws.merge_range('B12:B14',BB13, format)
            ws.write(11,2,BB14, format)
            ws.write(12,2,BB15, format)
            ws.write(13,2,BB16, format)
            ws.write(11,3,task.CPA22D1C, format)
            ws.write(12,3,task.CPA22D2C, format)
            ws.write(13,3,task.CPA22D3C, format)
            ws.write(11,4,task.CPA22D1A, format)
            ws.write(12,4,task.CPA22D2A, format)
            ws.write(13,4,task.CPA22D3A, format)
            ws.write(11,5,task.CPA22D1K, format)
            ws.write(12,5,task.CPA22D2K, format)
            ws.write(13,5,task.CPA22D3K, format)
            ws.write(11,6,task.CPA22D1P, format)
            ws.write(12,6,task.CPA22D2P, format)
            ws.write(13,6,task.CPA22D3P, format)
            ws.merge_range('B15:B17',BB17, format)
            ws.write(14,2,BB18, format)
            ws.write(15,2,BB19, format)
            ws.write(16,2,BB20, format)
            ws.write(14,3,task.CPA22E1C, format)
            ws.write(15,3,task.CPA22E2C, format)
            ws.write(16,3,task.CPA22E3C, format)
            ws.write(14,4,task.CPA22E1A, format)
            ws.write(15,4,task.CPA22E2A, format)
            ws.write(16,4,task.CPA22E3A, format)
            ws.write(14,5,task.CPA22E1K, format)
            ws.write(15,5,task.CPA22E2K, format)
            ws.write(16,5,task.CPA22E3K, format)
            ws.write(14,6,task.CPA22E1P, format)
            ws.write(15,6,task.CPA22E2P, format)
            ws.write(16,6,task.CPA22E3P, format)
            ws.write(1,1,'コンピテンシー', format)
            ws.write(1,2,'コンピテンシー詳細', format)
            ws.write(1,3,'本人評価', format)
            ws.write(1,4,'評定者評価', format)
            ws.write(1,5,'考察', format)
            ws.write(1,6,'向上プラン', format)
            ws.set_column('B:B',17)
            ws.set_column('C:C',50)
            ws.set_column('D:E',14)
            ws.set_column('F:G',50)
            book.close()
            output.seek(0)
            filename = 'competence.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response
        elif "send" in request.POST:        
            sbmt = 1
            task.CPA22C = sbmt
            task.save()


    params = {
        "FN":request.user.first_name,
        "LN":request.user.last_name,
        "data1":CPA22.objects.values_list('CPA22A1C','CPA22A2C','CPA22A3C','CPA22B1C','CPA22B2C','CPA22B3C','CPA22C1C','CPA22C2C','CPA22C3C','CPA22D1C','CPA22D2C','CPA22D3C','CPA22E1C','CPA22E2C','CPA22E3C').get(user=request.user),
        "data2":CPA22.objects.values_list('CPA22A1A','CPA22A2A','CPA22A3A','CPA22B1A','CPA22B2A','CPA22B3A','CPA22C1A','CPA22C2A','CPA22C3A','CPA22D1A','CPA22D2A','CPA22D3A','CPA22E1A','CPA22E2A','CPA22E3A').get(user=request.user),
        "data3":CPA22.objects.values_list('CPA22A1K','CPA22A2K','CPA22A3K','CPA22B1K','CPA22B2K','CPA22B3K','CPA22C1K','CPA22C2K','CPA22C3K','CPA22D1K','CPA22D2K','CPA22D3K','CPA22E1K','CPA22E2K','CPA22E3K').get(user=request.user),
        "data4":CPA22.objects.values_list('CPA22A1P','CPA22A2P','CPA22A3P','CPA22B1P','CPA22B2P','CPA22B3P','CPA22C1P','CPA22C2P','CPA22C3P','CPA22D1P','CPA22D2P','CPA22D3P','CPA22E1P','CPA22E2P','CPA22E3P').get(user=request.user),
        "B1":BB1,"B2":BB2,"B3":BB3,"B4":BB4,"B5":BB5,"B6":BB6,"B7":BB7,"B8":BB8,"B9":BB9,"B10":BB10,"B11":BB11,"B12":BB12,"B13":BB13,"B14":BB14,"B15":BB15,"B16":BB16,"B17":BB17,"B18":BB18,"B19":BB19,"B20":BB20,
        "sbmt":sbmt,
    }
    return render(request, 'MBO/CPA.html', params)


def CPA2(request,name):

    object_list = User.objects.all()
    for i in object_list:
        if i.username == name:
            ID = i.id
            FN = i.first_name
            LN = i.last_name
    
    AA=CPA22.objects.values_list('CPA22C').get(user=ID)
    CC=CPA22.objects.values_list('CPA22A').get(user=ID)
    if int(AA[0]) == 0:
        sbmt=0
    else:
        if int(CC[0]) == 0:
            sbmt = 1
        else:
            sbmt = 2
          
    task = get_object_or_404(CPA22, user=ID)
    if (request.method == 'POST'):
        if "DL" in request.POST:
            output = io.BytesIO()
            book = op.Workbook(output)
            book = XlsxWriter.Workbook(output)
            ws = book.add_worksheet('competency')
            format = book.add_format({'border':1})
            format.set_text_wrap()          
            ws.merge_range('B3:B5',BB1, format)
            ws.write(2,2,BB2, format)
            ws.write(3,2,BB3, format)
            ws.write(4,2,BB4, format)
            ws.write(2,3,task.CPA22A1C, format)
            ws.write(3,3,task.CPA22A2C, format)
            ws.write(4,3,task.CPA22A3C, format)
            ws.write(2,4,task.CPA22A1A, format)
            ws.write(3,4,task.CPA22A2A, format)
            ws.write(4,4,task.CPA22A3A, format)
            ws.write(2,5,task.CPA22A1K, format)
            ws.write(3,5,task.CPA22A2K, format)
            ws.write(4,5,task.CPA22A3K, format)
            ws.write(2,6,task.CPA22A1P, format)
            ws.write(3,6,task.CPA22A2P, format)
            ws.write(4,6,task.CPA22A3P, format)
            ws.merge_range('B6:B8',BB5, format)
            ws.write(5,2,BB6, format)
            ws.write(6,2,BB7, format)
            ws.write(7,2,BB8, format)
            ws.write(5,3,task.CPA22B1C, format)
            ws.write(6,3,task.CPA22B2C, format)
            ws.write(7,3,task.CPA22B3C, format)
            ws.write(5,4,task.CPA22B1A, format)
            ws.write(6,4,task.CPA22B2A, format)
            ws.write(7,4,task.CPA22B3A, format)
            ws.write(5,5,task.CPA22B1K, format)
            ws.write(6,5,task.CPA22B2K, format)
            ws.write(7,5,task.CPA22B3K, format)
            ws.write(5,6,task.CPA22B1P, format)
            ws.write(6,6,task.CPA22B2P, format)
            ws.write(7,6,task.CPA22B3P, format)
            ws.merge_range('B9:B11',BB9, format)
            ws.write(8,2,BB10, format)
            ws.write(9,2,BB11, format)
            ws.write(10,2,BB12, format)
            ws.write(8,3,task.CPA22C1C, format)
            ws.write(9,3,task.CPA22C2C, format)
            ws.write(10,3,task.CPA22C3C, format)
            ws.write(8,4,task.CPA22C1A, format)
            ws.write(9,4,task.CPA22C2A, format)
            ws.write(10,4,task.CPA22C3A, format)
            ws.write(8,5,task.CPA22C1K, format)
            ws.write(9,5,task.CPA22C2K, format)
            ws.write(10,5,task.CPA22C3K, format)
            ws.write(8,6,task.CPA22C1P, format)
            ws.write(9,6,task.CPA22C2P, format)
            ws.write(10,6,task.CPA22C3P, format)
            ws.merge_range('B12:B14',BB13, format)
            ws.write(11,2,BB14, format)
            ws.write(12,2,BB15, format)
            ws.write(13,2,BB16, format)
            ws.write(11,3,task.CPA22D1C, format)
            ws.write(12,3,task.CPA22D2C, format)
            ws.write(13,3,task.CPA22D3C, format)
            ws.write(11,4,task.CPA22D1A, format)
            ws.write(12,4,task.CPA22D2A, format)
            ws.write(13,4,task.CPA22D3A, format)
            ws.write(11,5,task.CPA22D1K, format)
            ws.write(12,5,task.CPA22D2K, format)
            ws.write(13,5,task.CPA22D3K, format)
            ws.write(11,6,task.CPA22D1P, format)
            ws.write(12,6,task.CPA22D2P, format)
            ws.write(13,6,task.CPA22D3P, format)
            ws.merge_range('B15:B17',BB17, format)
            ws.write(14,2,BB18, format)
            ws.write(15,2,BB19, format)
            ws.write(16,2,BB20, format)
            ws.write(14,3,task.CPA22E1C, format)
            ws.write(15,3,task.CPA22E2C, format)
            ws.write(16,3,task.CPA22E3C, format)
            ws.write(14,4,task.CPA22E1A, format)
            ws.write(15,4,task.CPA22E2A, format)
            ws.write(16,4,task.CPA22E3A, format)
            ws.write(14,5,task.CPA22E1K, format)
            ws.write(15,5,task.CPA22E2K, format)
            ws.write(16,5,task.CPA22E3K, format)
            ws.write(14,6,task.CPA22E1P, format)
            ws.write(15,6,task.CPA22E2P, format)
            ws.write(16,6,task.CPA22E3P, format)
            ws.write(1,1,'コンピテンシー', format)
            ws.write(1,2,'コンピテンシー詳細', format)
            ws.write(1,3,'本人評価', format)
            ws.write(1,4,'評定者評価', format)
            ws.write(1,5,'考察', format)
            ws.write(1,6,'向上プラン', format)
            ws.set_column('B:B',17)
            ws.set_column('C:C',50)
            ws.set_column('D:E',14)
            ws.set_column('F:G',50)
            book.close()
            output.seek(0)
            filename = 'competence.xlsx'
            response = HttpResponse(output, content_type='aplication/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s' %filename
            return response
        elif "back" in request.POST:        
            sbmt = 0
            task.CPA22C = sbmt
        elif "send" in request.POST:        
            sbmt = 2
            task.CPA22A = 1
            task.save()

    params = {
        "data1":CPA22.objects.values_list('CPA22A1C','CPA22A2C','CPA22A3C','CPA22B1C','CPA22B2C','CPA22B3C','CPA22C1C','CPA22C2C','CPA22C3C','CPA22D1C','CPA22D2C','CPA22D3C','CPA22E1C','CPA22E2C','CPA22E3C').get(user=ID),
        "data2":CPA22.objects.values_list('CPA22A1A','CPA22A2A','CPA22A3A','CPA22B1A','CPA22B2A','CPA22B3A','CPA22C1A','CPA22C2A','CPA22C3A','CPA22D1A','CPA22D2A','CPA22D3A','CPA22E1A','CPA22E2A','CPA22E3A').get(user=ID),
        "data3":CPA22.objects.values_list('CPA22A1K','CPA22A2K','CPA22A3K','CPA22B1K','CPA22B2K','CPA22B3K','CPA22C1K','CPA22C2K','CPA22C3K','CPA22D1K','CPA22D2K','CPA22D3K','CPA22E1K','CPA22E2K','CPA22E3K').get(user=ID),
        "data4":CPA22.objects.values_list('CPA22A1P','CPA22A2P','CPA22A3P','CPA22B1P','CPA22B2P','CPA22B3P','CPA22C1P','CPA22C2P','CPA22C3P','CPA22D1P','CPA22D2P','CPA22D3P','CPA22E1P','CPA22E2P','CPA22E3P').get(user=ID),
        "B1":BB1,"B2":BB2,"B3":BB3,"B4":BB4,"B5":BB5,"B6":BB6,"B7":BB7,"B8":BB8,"B9":BB9,"B10":BB10,"B11":BB11,"B12":BB12,"B13":BB13,"B14":BB14,"B15":BB15,"B16":BB16,"B17":BB17,"B18":BB18,"B19":BB19,"B20":BB20,
        "sbmt":sbmt,
        "EMP" : name,
        "FN" : FN,
        "LN" : LN,
        "FN2" : request.user.first_name,
        "LN2" : request.user.last_name,
    }
    return render(request, 'MBO/CPA2.html', params)


def edit(request, num):
    obj = GOAL22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = GOAL22Q1Form(request.POST, instance=obj)
        friend.save()
        return redirect(to='/GOAL')

    params = {
        "UserID":request.user,
        "data1":GOAL22Q1Form(instance=obj),
        "data2":GOAL22.objects.values_list('GOAL22A2','GOAL22B2','GOAL22C2','GOAL22D2','GOAL22E2','GOAL22F2','GOAL22G2').get(user=request.user),
        "data3":GOAL22.objects.values_list('GOAL22A3','GOAL22B3','GOAL22C3','GOAL22D3','GOAL22E3','GOAL22F3','GOAL22G3').get(user=request.user),
        "data4":GOAL22.objects.values_list('GOAL22AR','GOAL22BR','GOAL22CR','GOAL22DR','GOAL22ER','GOAL22FR','GOAL22GR').get(user=request.user),
        "data5":GOAL22.objects.values_list('GOAL22AJ','GOAL22BJ','GOAL22CJ','GOAL22DJ','GOAL22EJ','GOAL22FJ','GOAL22GJ').get(user=request.user),
        "FN":request.user.first_name,
        "LN":request.user.last_name,
        }
    return render(request, 'MBO/edit.html', params)

def editCPA(request, num):
    obj = CPA22.objects.get(user=request.user)
    if (request.method == 'POST'):
        friend = CPA22CForm(request.POST, instance=obj)
        friend.save()
        return redirect(to='/CPA')

    params = {
        "FN":request.user.first_name,
        "LN":request.user.last_name,
        "data1":CPA22CForm(instance=obj),
        "data2":CPA22.objects.values_list('CPA22A1A','CPA22A2A','CPA22A3A','CPA22B1A','CPA22B2A','CPA22B3A','CPA22C1A','CPA22C2A','CPA22C3A','CPA22D1A','CPA22D2A','CPA22D3A','CPA22E1A','CPA22E2A','CPA22E3A').get(user=request.user),
        "B1":BB1,"B2":BB2,"B3":BB3,"B4":BB4,"B5":BB5,"B6":BB6,"B7":BB7,"B8":BB8,"B9":BB9,"B10":BB10,"B11":BB11,"B12":BB12,"B13":BB13,"B14":BB14,"B15":BB15,"B16":BB16,"B17":BB17,"B18":BB18,"B19":BB19,"B20":BB20,
        }
    return render(request, 'MBO/editCPA.html', params)

def editCPA2(request,name):

    object_list = User.objects.all()
    for i in object_list:
        if i.username == name:
            ID = i.id
            FN = i.first_name
            LN = i.last_name

    obj = CPA22.objects.get(user=ID)    
    if (request.method == 'POST'):
        friend = CPA22AForm(request.POST, instance=obj)
        friend.save()
        return redirect('CPA2', name=name)

    params = {
        "data1":CPA22.objects.values_list('CPA22A1C','CPA22A2C','CPA22A3C','CPA22B1C','CPA22B2C','CPA22B3C','CPA22C1C','CPA22C2C','CPA22C3C','CPA22D1C','CPA22D2C','CPA22D3C','CPA22E1C','CPA22E2C','CPA22E3C').get(user=ID),
        "data2":CPA22AForm(instance=obj),
        "data3":CPA22.objects.values_list('CPA22A1K','CPA22A2K','CPA22A3K','CPA22B1K','CPA22B2K','CPA22B3K','CPA22C1K','CPA22C2K','CPA22C3K','CPA22D1K','CPA22D2K','CPA22D3K','CPA22E1K','CPA22E2K','CPA22E3K').get(user=ID),
        "data4":CPA22.objects.values_list('CPA22A1P','CPA22A2P','CPA22A3P','CPA22B1P','CPA22B2P','CPA22B3P','CPA22C1P','CPA22C2P','CPA22C3P','CPA22D1P','CPA22D2P','CPA22D3P','CPA22E1P','CPA22E2P','CPA22E3P').get(user=ID),
        "B1":BB1,"B2":BB2,"B3":BB3,"B4":BB4,"B5":BB5,"B6":BB6,"B7":BB7,"B8":BB8,"B9":BB9,"B10":BB10,"B11":BB11,"B12":BB12,"B13":BB13,"B14":BB14,"B15":BB15,"B16":BB16,"B17":BB17,"B18":BB18,"B19":BB19,"B20":BB20,
        "EMP" : name,
        "FN" : FN,
        "LN" : LN,
        "FN2" : request.user.first_name,
        "LN2" : request.user.last_name,
    }
    return render(request, 'MBO/editCPA2.html', params)


def homeA(request):
    object_list = User.objects.all()
    params = {
        "data" : object_list,
        "ID" : request.user.id
        }
    return render(request, 'MBO/homeA.html', params)

def homeB(request):
    object_list = User.objects.all()
    params = {
        "data" : object_list,
        "ID" : request.user.id
        }
    return render(request, 'MBO/homeB.html', params)


